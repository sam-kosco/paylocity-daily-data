"""
schedule_hours_sync.py
======================
Repo: github.com/sam-kosco/paylocity-daily-data

Nightly sync of Paylocity scheduling and punch data into Schedule_and_Hours.xlsx on SharePoint.

Schedule sheet  : 2 weeks forward (continuously updated — adds new shifts, updates changed ones,
                  removes cancelled ones). Also back-fills/updates the prior 2 weeks.
Labor Hours sheet: 60 days back. Each row = one employee on one date (all punches aggregated).
                   Only the last 14 days are re-queried each run to save API calls.

Run via GitHub Actions on a nightly cron, or manually:
    python schedule_hours_sync.py

Requires secrets (set in GitHub → Settings → Secrets → Actions for THIS repo):
    PAYLOCITY_CLIENT_ID      — Paylocity API client ID
    PAYLOCITY_CLIENT_SECRET  — Paylocity API client secret
    TENANT_ID                — Microsoft Entra tenant ID (ede0c57f-549f-4a90-9f8c-7ea130346f95)
    CLIENT_ID                — Entra app client ID for Foxtrot Report Automation
                               (58191600-ab56-4141-bff6-806805fcbff4)
    CLIENT_SECRET            — Entra app client secret (same one used in ramp-refresh and
                               stl-report-automation — renew in Entra, update ALL three repos)

NOTE: These secrets are NOT shared automatically between repos. You must add them manually
to paylocity-daily-data even though they already exist in ramp-refresh and stl-report-automation.
"""

import os
import io
import time
import requests
import pandas as pd
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo

# ── Config ────────────────────────────────────────────────────────────────────

PAYLOCITY_TOKEN_URL  = "https://api.paylocity.com/IdentityServer/connect/token"
PAYLOCITY_BASE       = "https://api.paylocity.com/api"
COMPANY_ID           = "350673"

GRAPH_TOKEN_URL      = "https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
GRAPH_BASE           = "https://graph.microsoft.com/v1.0"

# SharePoint paths
SP_SITE_ID           = "foxtrotaviationcom.sharepoint.com:/sites/DataHub:"
SP_EMPLOYEES_PATH    = "/Definitive Lists/Current Employees.csv"
SP_WORKBOOK_PATH     = "/Definitive Lists/Schedule and Hours.xlsx"

EST = ZoneInfo("America/New_York")

# ── Auth helpers ───────────────────────────────────────────────────────────────

def get_paylocity_token():
    import base64
    client_id     = os.environ["PAYLOCITY_CLIENT_ID"]
    client_secret = os.environ["PAYLOCITY_CLIENT_SECRET"]
    credentials   = base64.b64encode(f"{client_id}:{client_secret}".encode()).decode()
    resp = requests.post(
        PAYLOCITY_TOKEN_URL,
        data={"grant_type": "client_credentials"},
        headers={
            "Content-Type":  "application/x-www-form-urlencoded",
            "Authorization": f"Basic {credentials}",
        },
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def get_graph_token():
    url = GRAPH_TOKEN_URL.format(tenant=os.environ["TENANT_ID"])
    resp = requests.post(
        url,
        data={
            "grant_type":    "client_credentials",
            "client_id":     os.environ["CLIENT_ID"],
            "client_secret": os.environ["CLIENT_SECRET"],
            "scope":         "https://graph.microsoft.com/.default",
        },
        headers={"Content-Type": "application/x-www-form-urlencoded"},
    )
    resp.raise_for_status()
    return resp.json()["access_token"]

# ── Paylocity API helpers ──────────────────────────────────────────────────────

def paylocity_get(token, path, params=None, retries=3):
    headers = {"Authorization": f"Bearer {token}"}
    for attempt in range(retries):
        resp = requests.get(f"{PAYLOCITY_BASE}{path}", headers=headers, params=params)
        if resp.status_code == 401:
            raise RuntimeError("Paylocity token expired mid-run")
        if resp.status_code == 429:
            time.sleep(10 * (attempt + 1))
            continue
        resp.raise_for_status()
        return resp.json()
    raise RuntimeError(f"Failed after {retries} retries: {path}")


def get_shifts_for_employee(token, emp_id, start_dt, end_dt):
    """Return list of shift dicts for one employee over a date range."""
    path = f"/v2/companies/{COMPANY_ID}/employees/{emp_id}/scheduling/shifts"
    params = {
        "filter": f"startDateTime ge {start_dt.isoformat()} and startDateTime le {end_dt.isoformat()}",
        "limit":  200,
    }
    try:
        data = paylocity_get(token, path, params)
        return data if isinstance(data, list) else data.get("shifts", [])
    except Exception:
        return []


def get_punch_details_for_employee(token, emp_id, relative_start, relative_end):
    """
    Return aggregated punch data for one employee.
    relative_start / relative_end: ISO datetime strings in EST.
    Returns list of {date, hours_worked, labor_dist, work_scope, earnings}.
    """
    path = f"/apiHub/time/v2/companies/{COMPANY_ID}/employees/{emp_id}/punchdetails"
    params = {
        "relativeStart": relative_start,
        "relativeEnd":   relative_end,
    }
    try:
        data = paylocity_get(token, path, params)
        punches = data if isinstance(data, list) else data.get("punchDetails", [])
    except Exception:
        return []

    # Aggregate by date — only Clock Out punches carry duration/earnings
    by_date = {}
    for p in punches:
        if p.get("punchType") not in ("ClockOut", "Clock Out"):
            continue
        raw_date = p.get("punchDate", "")[:10]
        if not raw_date:
            continue
        if raw_date not in by_date:
            by_date[raw_date] = {
                "date":          raw_date,
                "hours_worked":  0.0,
                "labor_dist":    "",
                "work_scope":    "",
                "earnings":      0.0,
            }
        duration_mins = p.get("duration") or 0
        by_date[raw_date]["hours_worked"] += round(duration_mins / 60, 4)
        by_date[raw_date]["earnings"]     += p.get("earnings") or 0

        # Use cost center from first punch that has one
        ccs = p.get("costCenters", [])
        for cc in ccs:
            level = cc.get("level")
            code  = cc.get("code", "")
            if level == 0 and not by_date[raw_date]["labor_dist"]:
                by_date[raw_date]["labor_dist"] = code
            if level == 1 and not by_date[raw_date]["work_scope"]:
                by_date[raw_date]["work_scope"] = code

    return list(by_date.values())

# ── Graph / SharePoint helpers ─────────────────────────────────────────────────

def get_drive_id(graph_token):
    headers = {"Authorization": f"Bearer {graph_token}"}
    resp = requests.get(
        f"{GRAPH_BASE}/sites/{SP_SITE_ID}/drive",
        headers=headers,
    )
    resp.raise_for_status()
    return resp.json()["id"]


def download_file(graph_token, drive_id, path):
    headers = {"Authorization": f"Bearer {graph_token}"}
    resp = requests.get(
        f"{GRAPH_BASE}/drives/{drive_id}/root:{path}:/content",
        headers=headers,
    )
    resp.raise_for_status()
    return resp.content


def upload_file(graph_token, drive_id, path, content_bytes):
    headers = {
        "Authorization": f"Bearer {graph_token}",
        "Content-Type":  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    resp = requests.put(
        f"{GRAPH_BASE}/drives/{drive_id}/root:{path}:/content",
        headers=headers,
        data=content_bytes,
    )
    resp.raise_for_status()
    return resp.json()

# ── Main sync logic ────────────────────────────────────────────────────────────

def main():
    now_est   = datetime.now(EST)
    today     = now_est.date()

    # Date windows
    sched_start = today - timedelta(days=14)      # 2 weeks back (update window)
    sched_end   = today + timedelta(days=14)      # 2 weeks forward
    labor_full_start = today - timedelta(days=60) # 60 days back (full history)
    labor_update_start = today - timedelta(days=14) # only re-query last 14 days

    print(f"Run date (EST): {today}")
    print(f"Schedule window: {sched_start} → {sched_end}")
    print(f"Labor update window: {labor_update_start} → {today}")
    print(f"Labor full history: {labor_full_start} → {today}")

    # ── Tokens ────────────────────────────────────────────────────────────────
    print("\nAuthenticating...")
    pcty_token  = get_paylocity_token()
    graph_token = get_graph_token()

    # ── Get employee list from SharePoint CSV ─────────────────────────────────
    print("Downloading employee list from SharePoint...")
    drive_id = get_drive_id(graph_token)
    emp_csv_bytes = download_file(graph_token, drive_id, SP_EMPLOYEES_PATH)
    emp_df = pd.read_csv(
        io.BytesIO(emp_csv_bytes),
        dtype={"Employee Id": str, "Supervisor ID": str},
    )
    # Filter to Active only
    emp_df = emp_df[emp_df["Status"] == "Active"].copy()
    emp_df["Employee Id"] = emp_df["Employee Id"].str.strip()

    # Build name lookup: id → "First Last"
    emp_df["Full Name"] = emp_df["First Name"].str.strip() + " " + emp_df["Last Name"].str.strip()
    name_lookup = emp_df.set_index("Employee Id")["Full Name"].to_dict()

    employee_ids = emp_df["Employee Id"].tolist()
    print(f"  {len(employee_ids)} active employees found")

    # ── Download existing workbook ─────────────────────────────────────────────
    print("Downloading existing workbook from SharePoint...")
    wb_bytes = download_file(graph_token, drive_id, SP_WORKBOOK_PATH)
    existing_sheets = pd.read_excel(io.BytesIO(wb_bytes), sheet_name=None, dtype=str)

    # Load existing Schedule data
    sched_df = existing_sheets.get("Schedule", pd.DataFrame())
    if sched_df.empty or "Shift ID" not in sched_df.columns:
        sched_df = pd.DataFrame(columns=[
            "Shift ID", "Employee ID", "Employee Name",
            "Start DateTime", "End DateTime",
            "Labor Distribution", "Work Scope", "Scheduled Time",
        ])

    # Load existing Labor Hours data
    labor_df = existing_sheets.get("Labor Hours", pd.DataFrame())
    if labor_df.empty or "Employee ID" not in labor_df.columns:
        labor_df = pd.DataFrame(columns=[
            "Employee ID", "Employee Name", "Date",
            "Hours Worked", "Labor Distribution", "Work Scope", "Earnings",
        ])

    # ── SCHEDULE: fetch and upsert ────────────────────────────────────────────
    print(f"\nFetching schedule ({sched_start} → {sched_end})...")

    sched_start_dt = datetime(sched_start.year, sched_start.month, sched_start.day,
                               tzinfo=EST)
    sched_end_dt   = datetime(sched_end.year,   sched_end.month,   sched_end.day,
                               23, 59, 59, tzinfo=EST)

    new_shifts = []
    for i, emp_id in enumerate(employee_ids):
        if i % 50 == 0:
            print(f"  Schedule: {i}/{len(employee_ids)} employees...")
            # Refresh Paylocity token every 50 employees (tokens last 1 hour)
            if i > 0:
                pcty_token = get_paylocity_token()

        shifts = get_shifts_for_employee(pcty_token, emp_id, sched_start_dt, sched_end_dt)
        for s in shifts:
            start_raw = s.get("startDateTime", "")
            duration_mins = s.get("duration", 0) or 0

            # Parse start, compute end
            try:
                start_dt = datetime.fromisoformat(start_raw).astimezone(EST)
                end_dt   = start_dt + timedelta(minutes=duration_mins)
                start_str = start_dt.strftime("%Y-%m-%d %H:%M")
                end_str   = end_dt.strftime("%Y-%m-%d %H:%M")
                sched_hrs = round(duration_mins / 60, 2)
            except Exception:
                start_str = start_raw
                end_str   = ""
                sched_hrs = round(duration_mins / 60, 2)

            # Cost centers
            labor_dist = ""
            work_scope = ""
            for cc in s.get("costCenters", []):
                if cc.get("level") == 0:
                    labor_dist = cc.get("code", "")
                if cc.get("level") == 1:
                    work_scope = cc.get("code", "")

            new_shifts.append({
                "Shift ID":          s.get("shiftId", ""),
                "Employee ID":       emp_id,
                "Employee Name":     name_lookup.get(emp_id, ""),
                "Start DateTime":    start_str,
                "End DateTime":      end_str,
                "Labor Distribution": labor_dist,
                "Work Scope":        work_scope,
                "Scheduled Time":    sched_hrs,
            })

    new_sched_df = pd.DataFrame(new_shifts)

    # Upsert: keep existing rows outside the update window, replace within it
    if not sched_df.empty and "Start DateTime" in sched_df.columns:
        # Keep rows whose start date is before the update window (historical)
        try:
            sched_df["_start_date"] = pd.to_datetime(
                sched_df["Start DateTime"], errors="coerce"
            ).dt.date
            outside_window = sched_df[sched_df["_start_date"] < sched_start].drop(
                columns=["_start_date"]
            )
        except Exception:
            outside_window = pd.DataFrame(columns=sched_df.columns)
    else:
        outside_window = pd.DataFrame(columns=[
            "Shift ID", "Employee ID", "Employee Name",
            "Start DateTime", "End DateTime",
            "Labor Distribution", "Work Scope", "Scheduled Time",
        ])

    final_sched = pd.concat([outside_window, new_sched_df], ignore_index=True)
    final_sched = final_sched.drop_duplicates(subset=["Shift ID"], keep="last")

    print(f"  Schedule rows: {len(final_sched)} total ({len(new_shifts)} from API)")

    # ── LABOR HOURS: fetch and upsert ─────────────────────────────────────────
    print(f"\nFetching labor hours ({labor_update_start} → {today})...")

    labor_start_dt = datetime(labor_update_start.year, labor_update_start.month,
                               labor_update_start.day, tzinfo=EST)
    labor_end_dt   = datetime(today.year, today.month, today.day,
                               23, 59, 59, tzinfo=EST)

    new_labor_rows = []
    for i, emp_id in enumerate(employee_ids):
        if i % 50 == 0:
            print(f"  Labor: {i}/{len(employee_ids)} employees...")
            if i > 0:
                pcty_token = get_paylocity_token()

        rows = get_punch_details_for_employee(
            pcty_token, emp_id,
            labor_start_dt.isoformat(),
            labor_end_dt.isoformat(),
        )
        for r in rows:
            new_labor_rows.append({
                "Employee ID":        emp_id,
                "Employee Name":      name_lookup.get(emp_id, ""),
                "Date":               r["date"],
                "Hours Worked":       round(r["hours_worked"], 2),
                "Labor Distribution": r["labor_dist"],
                "Work Scope":         r["work_scope"],
                "Earnings":           round(r["earnings"], 2),
            })

    new_labor_df = pd.DataFrame(new_labor_rows)

    # Upsert: keep existing rows outside the 14-day update window
    # (preserve the 60-day history that was already written in prior runs)
    if not labor_df.empty and "Date" in labor_df.columns:
        try:
            labor_df["_date"] = pd.to_datetime(labor_df["Date"], errors="coerce").dt.date
            # Keep rows older than update window AND not in new data
            outside_labor = labor_df[
                labor_df["_date"] < labor_update_start
            ].drop(columns=["_date"])
            # Also keep rows beyond 60 days only if they exist (shouldn't, but safe)
        except Exception:
            outside_labor = pd.DataFrame(columns=labor_df.columns)
    else:
        outside_labor = pd.DataFrame(columns=[
            "Employee ID", "Employee Name", "Date",
            "Hours Worked", "Labor Distribution", "Work Scope", "Earnings",
        ])

    final_labor = pd.concat([outside_labor, new_labor_df], ignore_index=True)
    # Deduplicate on Employee + Date
    final_labor = final_labor.drop_duplicates(
        subset=["Employee ID", "Date"], keep="last"
    )
    # Drop rows older than 60 days
    try:
        final_labor["_date"] = pd.to_datetime(final_labor["Date"], errors="coerce").dt.date
        final_labor = final_labor[
            final_labor["_date"] >= labor_full_start
        ].drop(columns=["_date"])
    except Exception:
        pass

    final_labor = final_labor.sort_values(["Date", "Employee ID"]).reset_index(drop=True)

    print(f"  Labor rows: {len(final_labor)} total ({len(new_labor_rows)} from API)")

    # ── Write workbook and upload ──────────────────────────────────────────────
    print("\nWriting workbook...")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))

    def write_sheet(wb, sheet_name, df, header_fill_hex="1B2D6B"):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(sheet_name)

        header_fill = PatternFill("solid", fgColor=header_fill_hex)
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        header_align = Alignment(horizontal="center", vertical="center")
        body_font   = Font(name="Calibri", size=10)

        # Write header
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font    = header_font
            cell.fill    = header_fill
            cell.alignment = header_align

        # Write data rows
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = body_font

        # Auto-size columns
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(
                len(str(col_name)),
                *(len(str(v)) for v in df.iloc[:, col_idx - 1] if pd.notna(v)),
                default=len(col_name),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

    write_sheet(wb, "Schedule",     final_sched,  header_fill_hex="1B2D6B")
    write_sheet(wb, "Labor Hours",  final_labor,  header_fill_hex="1A7A8A")

    # Serialise to bytes
    buf = io.BytesIO()
    wb.save(buf)
    wb_bytes_out = buf.getvalue()

    print("Uploading workbook to SharePoint...")
    upload_file(graph_token, drive_id, SP_WORKBOOK_PATH, wb_bytes_out)
    print("Done.")
    print(f"\nSummary:")
    print(f"  Schedule rows written : {len(final_sched)}")
    print(f"  Labor Hours rows written: {len(final_labor)}")
    print(f"  Run completed at: {datetime.now(EST).strftime('%Y-%m-%d %H:%M:%S')} EST")


if __name__ == "__main__":
    main()
